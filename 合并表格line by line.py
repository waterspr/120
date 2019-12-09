#  -*- coding :utf8 -*-

from  openpyxl import  Workbook    #导入库
from openpyxl  import load_workbook  #导入库
import os 

#创建工作簿和工作表
wb3 = Workbook()
# 激活 worksheet
ws3 = wb3.active
a = 1

for root, dirs, files in os.walk(r'd:\1'):   #os.walk()用法,历遍文件下的目录和文件
    for name in files :
        # print(os.path.join(root, name))
        ph = os.path.join(root, name)     #工作簿excel的路径地址
        print(ph)
        wb = load_workbook(ph)   #导入工作簿
        ws = wb.worksheets[0]    #导入工作表
        for i in range(1,ws.max_row+1) :      #历遍每一行
            for j in range(1,ws.max_column+1) :    #历遍每一列  
                ws3.cell(row=a, column=j).value = ws.cell(row=i, column=j).value    # 新表和旧表的value值对应写入
            a +=1   #等同 a = a +1                #新表ws3 换行写入

wb3.save(r'D:\001.xlsx')    #保存wb3工作簿
print('done')                



