# 1、 安装

# pip install openpyxl

# 想要在文件中插入图片文件，需要安装pillow，安装文件：PIL-fork-1.1.7.win-amd64-py2.7.exe


# · font(字体类)：字号、字体颜色、下划线等

# · fill(填充类)：颜色等

# · border(边框类)：设置单元格边框

# · alignment(位置类)：对齐方式

# · number_format(格式类)：数据格式

# · protection(保护类)：写保护

'''
####2、 创建一个excel 文件，并写入不同类的内容
# -*- coding : utf-8 -*-

from openpyxl import Workbook
wb = Workbook()   #创建文件对象
#grab the active worksheet
ws = wb.active   #获取第一个sheet
ws['A1']=45    #写入数字
ws['B1']='你好'+ '这是测试'   #写入文本
#rows can also be appended
ws.append([1,2,3])  #写入多个单元格
#python types will automaically be converted
import datetime
import time
ws['A2'] = datetime.datetime.now()  #写入一个当前时间
#写入一个自定义的时间格式
ws['A3'] =time.strftime('%Y{y}%m{m}%d{d} %H{h}%M{f}%S{s}',time.localtime()).format(y='年', m='月', d='日', h='时', f='分', s='秒')
#save the file 保存文件
wb.save('d:\\1\\新建测试表格.xlsx')
print ('完成')

'''

print('this is {},that is {}'.format('鸡','蛋'))


'''
###3、 创建sheet
# -*- coding: utf-8 -*-

from openpyxl import Workbook
wb = Workbook()
ws1 = wb.create_sheet('mysheet') #创建一个sheet
ws1.title = 'new title'           #设定一个sheet的名字
ws2 = wb.create_sheet('mysheet2',0) #设定sheet的插入位置 默认插在后面
ws2.title = u'表名'      #设定一个sheet的名字 必须是Unicode
ws1.sheet_properties.tabColor = "1072BA"   #设定sheet的标签的背景颜色

#获取某个sheet对象
print(wb.get_sheet_by_name(u'表名'))
print(wb['new title'])

#获取全部sheet 的名字，遍历sheet名字
print(wb.sheetnames)
for sheet_name in wb.sheetnames :
    print(sheet_name)
print("*"*30)
for sheet in wb :
    print(sheet.title)

'''

'''
###操作单元格
# -*- coding :utf-8  -*-
from openpyxl import Workbook
wb = Workbook()
ws1 = wb.create_sheet('new sheet')
ws1['A1'] = 123
ws1['B2'] = u'测试'
d = ws1.cell(row=3,column=3,value=5)

print(ws1['A1'])
print(ws1['B2'])
print(ws1['A1'].value)
print(ws1['B2'].value)
print (d.value)
print(ws1["C3"].value)

wb.save("d:\\1\\new excel.xlsx")

'''

'''
###操作批量的单元格
# 无论ws.rows还是ws.iter_rows都是一个对象
# 除上述两个对象外 单行，单列都是一个元祖，多行多列是二维元祖

# -*- coding: utf-8 -*-
from openpyxl import Workbook
wb = Workbook()
ws1 = wb.active  #获取一个活动表
ws1["A1"]=1
ws1["A2"]=2
ws1["A3"]=3

ws1["B1"]=4
ws1["B2"]=5
ws1["B3"]=6

ws1["C1"]=7
ws1["C2"]=8
ws1["C3"]=9

#操作单列
print (ws1['A'])
for cell in ws1["A"] :
    print(cell.value)

#操作多列，获取每一个值
print(ws1['A:C'])
for column in ws1['A:C'] :
    for cell in column :
        print(cell.value)

#操作单行
print(ws1['1'])

#操作多行
row_range = ws1[1:3]
print(row_range)
for row in row_range :
    for cell in row :
        print(cell.value)

#获取所有行
print (ws1.rows)
for row in ws1.rows :
    print(row)
    for cell in row :
        print(cell.value)

#获取所有列
print(ws1.columns)
for col in ws1.columns :
    print (col)
    for cell in col :
        print(cell.value)

'''

'''
###获取所有的行对象：
#coding=utf-8
from openpyxl import Workbook
from openpyxl import load_workbook
wb = load_workbook(r'd:\1\new excel.xlsx')
ws=wb.active
rows=[]
for row in ws.iter_rows():
            rows.append(row)

print(rows)   #所有行
print(rows[0])   #获取第一行
print (rows[0][0])  #获取第一行第一列单元格对象
print (rows[0][0].value)  #获取第一行第一列单元格对象的值

print( rows[len(rows)-1] ) #获取最后行 print rows[-1]
print( rows[len(rows)-1][len(rows[0])-1] ) #获取第后一行和最后一列的单元格对象
print( rows[len(rows)-1][len(rows[0])-1].value ) #获取第后一行和最后一列的单元格对象的值

'''

'''
###获取所有的列对象：
#coding=utf-8
from openpyxl import Workbook
from openpyxl import load_workbook
wb = load_workbook('d:\\1\\new excel.xlsx')
ws=wb.active
cols=[]
for col in ws.iter_cols():
    cols.append(col)
 
print ( cols  ) #所有列
print (cols[0]  ) #获取第一列
print (cols[0][0] )  #获取第一列的第一行的单元格对象
print (cols[0][0].value)   #获取第一列的第一行的值

print ("*"*30)
print (cols[len(cols)-1])   #获取最后一列
print (cols[len(cols)-1][len(cols[0])-1])   #获取最后一列的最后一行的单元格对象
print (cols[len(cols)-1][len(cols[0])-1].value  ) #获取最后一列的最后一行的单元格对象的值

'''


###6、 操作已经存在的文件
# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl import load_workbook
wb = load_workbook('d:\\1\\new excel.xlsx')
wb.guess_types = True   #猜测格式类型
ws=wb.active
ws["D4"]="12%"
print (ws["D4"].value)
# Save the file
wb.save("d:\\1\\new excel.xlsx")
#注意如果原文件有一些图片或者图标，则保存的时候可能会导致图片丢失