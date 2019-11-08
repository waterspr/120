# -*- coding:utf-8 -*-
# message 处理数据

from openpyxl import load_workbook
import os
from openpyxl import Workbook
# from openpyxl.drawing.image import Image  # 注意当你想往excel写入图片时，可使用这个方法。


'''
def find_excel(path):  #  获取excel文档内容的方法
    wb = load_workbook(path)
    content = []
    wb.guess_types = True   #猜测格式类型
    sheet = wb.active
    for col in sheet.columns:
        for cell in col:   #  这个迭代无特殊情况 可简化一下
            # print(cell.value)
            content.append(cell.value)    #  注意，获取到的是一行的值

    return content

'''
def find_excel(path):  #  获取excel文档内容的方法
    wb = load_workbook(path)
    content = []
    wb.guess_types = True   #猜测格式类型
    sheet = wb.active
    for index, row in enumerate(sheet.rows):
         if index > 0:  # 因为 index = 0时，获取到的为表格的值。
             # print(row)
             for item in row:   #  这个迭代无特殊情况 可简化一下
                #  print(item.value)
                content.append(item.value)    #  注意，获取到的是一行的值
        
    return content

wb = Workbook()
ws1 = wb.create_sheet('银行',0) #创造合并之后的sheet表的名字
title_list = ['名字','个人照片', '二维码']  # sheet表的表头创建
ws1.append(title_list)

for root, dirs, files in os.walk('d:\\1'):  # 获取这个文件夹下所有的excel文档。
    for file in files:
        path_list = file   #os.walk()用法导致的,迭代遍历所有excel 文档
        print('excel文件名是%s' % path_list)
        #ph = ph.replace('~$', '')  #replace替换字符串的
        #if str(ph).endswith('sx'):  #endswith判断字符串是否以指定后缀结尾，如果以指定后缀结尾返回True，否则返回False。
        print (os.path.join(root, path_list))
        res = find_excel(os.path.join(root, path_list))   # 构建单个文档路径
        ws1.append(res)
        wb.save(r'D:\FX.xlsx')


print('done!!!')


'''
import os
from openpyxl import Workbook
from openpyxl import load_workbook



path = 'd:\\1\\1.xlsx'
wb = load_workbook(path)
wb.guess_types = True   #猜测格式类型
sheet = wb.active
for row in sheet.rows:
    for cell in row:   #  这个迭代无特殊情况 可简化一下
        print(cell.value)
'''
