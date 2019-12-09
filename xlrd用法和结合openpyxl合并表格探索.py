
#  -*- coding :utf8 -*-
import xlrd

###打开excel文件并获取所有sheet
workbook = xlrd.open_workbook(r'D:\1\123.xlsx')  #data = xlrd.open_workbook(filename)#文件名以及路径，如果路径或者文件名有中文给前面加一个r拜师原生字符。
'''
print(workbook.sheet_names())  # 和 print(list(names)) 功能一样

####获取book中一个工作表
#第一方法： 通过索引顺序获取 对象表
table1 = workbook.sheets()[0]  #第一方法： 通过索引顺序获取 对象表
print(table1)   #对象sheet1 
print(table1.name)  #  第一个sheet的名称

##获取工作表的所有名称
names = workbook.sheet_names()
print(list(names))   #输出工作表的名字，是列表形式
table1_name = workbook.sheet_names()[0] #names = data.sheet_names()返回book中所有工作表的名字
print(table1_name)  #  第一个sheet的名称



#第二方法： 通过索引顺序获取 对象表
table2 = workbook.sheet_by_index(0)   #第二方法： 通过索引顺序获取 对象表
print(table2)
print (workbook.sheet_names()[1])  #  第二个sheet的名称


#第三种方法： 通过名称获取， 注意名称的大小写
table3 = workbook.sheet_by_name('Sheet1')  #第三种方法： 通过名称获取， 注意名称的大小写
print(table3.name)

'''

'''
##(2). 行(line)的操作
table = workbook.sheet_by_index(0)
nrows = table.nrows   #获取该sheet中的有效行数
print(nrows)
row1 = table.row(0)  #返回由该行中所有的单元格 对象 组成的列表
print(row1)
row2type = table.row_types(1, start_colx=0, end_colx=None)    #返回由该行中所有单元格的 数据类型 组成的列表
print(row2type)
#type类型说明：ctype : 0 empty,1 string, 2 number, 3 date, 4 boolean, 5 error

# row2 = table.row_values(1, start_colx=0, end_colx=None)   #返回由该行中所有单元格的数据组成的列表
row2=table.row_values(1)   # start_colx=0, end_colx=None可省略
print(row2)
row1_len = table.row_len(0) #返回该行的有效单元格长度
print(row1_len)
'''
'''
###(3). 列(colnum)的操作
table = workbook.sheet_by_index(0)
ncols = table.ncols   #获取列表的有效列数
print(ncols)
col2 = table.col(1, start_rowx=0, end_rowx=None)  #返回由该列中所有的单元格对象组成的列表
print(col2)
col2_type = table.col_types(1, start_rowx=0, end_rowx=None)    #返回由该列中所有单元格的数据类型组成的列表
print(col2_type)
col1 = table.col_values(0, start_rowx=0, end_rowx=None)   #返回由该列中所有单元格的数据组成的列表
print(col1)
'''
'''
### (4). 单元格的操作
table = workbook.sheet_by_index(0)
a = table.cell(1,1)  #返回单元格对象,,第二行第二列 对象
print(a)
a_type = table.cell_type(1,1)    #返回单元格中的数据类型
print(a_type)
a_value = table.cell_value(1,1)   #返回单元格中的数据
print(a_value)
## 可以用encode('utf-8') 输出中文 sheet2.cell(1,0).value.encode('utf-8')
'''

'''
from  openpyxl import  Workbook 
from openpyxl  import load_workbook

wb = load_workbook(r'D:\1\123.xlsx')
# 激活 worksheet
#ws = wb.active
ws = wb.worksheets[1]
'''

'''
2.Workbook对象属性（工作簿操作）
 sheetnames：获取工作簿中的表（列表）
 active：获取当前活跃的Worksheet
 worksheets：以列表的形式返回所有的Worksheet(表格)
 read_only：判断是否以read_only模式打开Excel文档
 encoding：获取文档的字符集编码
 properties：获取文档的元数据，如标题，创建者，创建日期等

'''

'''
#openpyxl 获取工作表 ，给单元格赋值和输出
# sheet 名称可以作为 key 进行索引
     #ws3 = wb["New Title"]
     # ws4 = wb.get_sheet_by_name("New Title")
print(ws['A1'])
ws['A1'] = '99'    #赋值  方式一
ws.cell(row=2, column=2, value=10)    #赋值  方式二
ws.append([1, 2, 3])    #赋值 方式三：可以附加行，从第一列开始附加(从最下方空白处，最左开始)(可以输入多行)

print(ws['A1'].value)    #输出单元格中值
print(ws.cell(row=2, column=2).value)  #输出单元格中值
'''

from  openpyxl import  Workbook 
from openpyxl  import load_workbook

wb1 = xlrd.open_workbook(r'D:\1\123.xlsx')
wb2 = xlrd.open_workbook(r'D:\1\234.xlsx')
ws1 = wb1.sheet_by_index(1)
ws2 = wb2.sheet_by_index(0)
print (ws1.name ,ws2.name) #打印两个工作表的名称

#创建工作簿和工作表
wb3 = Workbook()
# 激活 worksheet
ws3 = wb3.active

nrows1 = ws1.nrows  
print(nrows1)
# print (ws1.row_values(1,start_colx=0, end_colx=None))
for row1 in range(0,nrows1+1) :
    try :
        data1 = ws1.row_values(row1,start_colx=0, end_colx=None)
        print(data1)
    except:
        continue
    ws3.append(data1)

nrows2 = ws2.nrows 
for row2 in range(0,nrows2+1) :
    try :
        data2 = ws2.row_values(row2,start_colx=0, end_colx=None)
        print(data2)
    except:
        continue
    ws3.append(data2)
    
# '''
# 说明 
#      这个地方加try 和 except 的原因是：
#     File "D:\ProgramData\Anaconda3\lib\site-packages\xlrd\sheet.py", line 493, in row_values
#     return self._cell_values[rowx][start_colx:end_colx]
#   IndexError: list index out of range
#   导致的原因有：
#     情况一：
#   list[index]中的index下标超出范围了，所以出现了访问越界；
#     情况二：
#   list本身就是一个空的，没有一个元素，所以当访问到list[0]的时候，就会出现该错误。

#   这个地方的原因是情况二：
#   查看row_values方法的源码：

#   def row_values(self, rowx, start_colx=0, end_colx=None):
#         if end_colx is None:
#             return self._cell_values[rowx][start_colx:]
#         return self._cell_values[rowx][start_colx:end_colx]

# '''



# print (ws1.row_values(1))


wb3.save(r'D:\1\00.xlsx')
print('done')


